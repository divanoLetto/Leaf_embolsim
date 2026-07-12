"""
evaluate.py — Quantitative evaluation of predicted incremental masks.



Usage (from project root):
    python src/method/evaluate.py --sequence Senecio_17_11_L5_Cavicam12_210725
    python src/method/evaluate.py --all
    python src/method/evaluate.py --sequence ... --vis-every 5
"""

import argparse
import logging
import os
import sys

import numpy as np
from PIL import Image
from scipy.ndimage import label as _cc_label

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from method import config
from method.dataset import _list_frames, _list_masks, _load_rgb, _load_mask, find_sequences


def _particles_area(mask: np.ndarray, min_size: int = 80) -> int:
    """Area (px) of connected components >= min_size — mimics ImageJ Analyze Particles (80-infinity)."""
    if mask.sum() == 0:
        return 0
    lab, n = _cc_label(mask)
    if n == 0:
        return 0
    sizes = np.bincount(lab.ravel())
    sizes[0] = 0
    return int(sizes[sizes >= min_size].sum())


_EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "data", "SenecioIVERdroughtVCs.xlsx",
)


def _excel_wp_series(seq_name: str):
    """Return (image_no_list, wp_corrected_list) for the given sequence, or (None, None).
    Lists are aligned by row position in the Excel sheet; image_no is 1-based."""
    import re
    m = re.match(r"Senecio_(\d+_\d+)_L(\d+)_", seq_name)
    if not m:
        return None, None
    sheet_name, leaf_n = m.group(1), int(m.group(2))
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_EXCEL_PATH, data_only=True, read_only=True)
    except Exception:
        return None, None
    if sheet_name not in wb.sheetnames:
        for s in wb.sheetnames:
            if s.startswith(sheet_name):
                sheet_name = s
                break
        else:
            return None, None
    ws = wb[sheet_name]
    col = 1 + 10 * (leaf_n - 1)
    hdr = ws.cell(row=1, column=col).value
    if hdr is None or "plant" not in str(hdr).lower():
        return None, None
    imgs, wps = [], []
    for r in range(2, ws.max_row + 1):
        img = ws.cell(row=r, column=col + 1).value
        wp  = ws.cell(row=r, column=col + 7).value  # 'wp corrected'
        if img is None:
            break
        if isinstance(img, (int, float)) and isinstance(wp, (int, float)):
            imgs.append(int(img))
            wps.append(float(wp))
    if not imgs:
        return None, None
    return imgs, wps


def _save_fp50_x_wp(pred_cum, gt_cum, seq_name, out_path):
    """Plot cumulative embolism (%) vs water potential (MPa) on the x-axis.
    Vertical lines mark fp50 for pred / GT (mask) / GT (Excel), each annotated with the wp value."""
    try:
        import matplotlib.pyplot as plt
    except ImportError:
        return
    imgs, wps = _excel_wp_series(seq_name)
    if imgs is None:
        return  # cannot draw without wp axis

    # Map frame_index (0-based) -> wp via image_no = idx + 1
    img_to_wp = dict(zip(imgs, wps))
    n = min(len(pred_cum), len(gt_cum))
    x_wp, y_pred, y_gt = [], [], []
    pred_total = pred_cum[-1] if pred_cum else 0
    gt_total   = gt_cum[-1]   if gt_cum   else 0
    for i in range(n):
        wp = img_to_wp.get(i + 1)
        if wp is None:
            continue
        x_wp.append(wp)
        y_pred.append(pred_cum[i] / pred_total * 100 if pred_total > 0 else 0)
        y_gt.append(  gt_cum[i]   / gt_total   * 100 if gt_total   > 0 else 0)
    if not x_wp:
        return

    pred_fp50 = _fp50_frame(pred_cum)
    gt_fp50   = _fp50_frame(gt_cum)
    ex_idx, ex_imgno = _excel_fp50_frame(seq_name)

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(x_wp, y_pred, label="Predicted",     lw=1.5)
    ax.plot(x_wp, y_gt,   label="GT (mask)",     lw=1.5, ls="--")

    def _annot(frame_idx, color, ls, label_prefix):
        if frame_idx is None or frame_idx < 0:
            return
        wp = img_to_wp.get(frame_idx + 1)
        if wp is None:
            return
        ax.axvline(wp, color=color, ls=ls, lw=1.5,
                   label=f"{label_prefix} (wp={wp:.2f} MPa, frame {frame_idx+1})")
        ax.text(wp, 102, f"{wp:.2f} MPa", color=color, rotation=90,
                ha="right", va="top", fontsize=8)

    _annot(pred_fp50, "C0", ":",  "fp50 pred")
    _annot(gt_fp50,   "C1", ":",  "fp50 GT mask")
    if ex_idx is not None:
        _annot(ex_idx, "C2", "-.", "fp50 GT Excel")

    ax.invert_xaxis()  # wp becomes more negative over time → time flows left-to-right
    ax.set_xlabel("Water potential, wp corrected (MPa)")
    ax.set_ylabel("% of embolism")
    ax.set_title(f"FP50 vs water potential — {seq_name}")
    ax.set_ylim(0, 105)
    ax.legend(loc="upper left", fontsize=8)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path, dpi=120)
    plt.close(fig)


def _excel_fp50_frame(seq_name: str):
    """Return (frame_index_0based, image_no) of first row with %emb >= 50 in the
    Excel sheet matching this sequence, or (None, None) if not found.

    Sequence naming convention: Senecio_<sheet>_L<leaf_n>_...  e.g. Senecio_10_11_L4_...
    Excel layout: each leaf occupies 10 columns starting at col 1 (plant ID), 11, 21, 31, 41.
    Within a block: plant_id (offset 0), image no (1), area (2), cumm emb (3), % emb (4).
    """
    import re
    m = re.match(r"Senecio_(\d+_\d+)_L(\d+)_", seq_name)
    if not m:
        return None, None
    sheet_name, leaf_n = m.group(1), int(m.group(2))
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_EXCEL_PATH, data_only=True, read_only=True)
    except Exception:
        return None, None
    if sheet_name not in wb.sheetnames:
        for s in wb.sheetnames:
            if s.startswith(sheet_name):
                sheet_name = s
                break
        else:
            return None, None
    ws = wb[sheet_name]
    col = 1 + 10 * (leaf_n - 1)  # plant ID column for this leaf
    hdr = ws.cell(row=1, column=col).value
    if hdr is None or "plant" not in str(hdr).lower():
        return None, None
    for r in range(2, ws.max_row + 1):
        img = ws.cell(row=r, column=col + 1).value
        pct = ws.cell(row=r, column=col + 4).value
        if img is None:
            break
        if isinstance(pct, (int, float)) and pct >= 50:
            try:
                return int(img) - 1, int(img)  # convert 1-based image_no to 0-based index
            except (TypeError, ValueError):
                return None, None
    return None, None


# ---------------------------------------------------------------------------
# Metric helpers
# ---------------------------------------------------------------------------

def pixel_metrics(pred: np.ndarray, gt: np.ndarray) -> dict:
    TP = int(np.logical_and(pred, gt).sum())
    FP = int(np.logical_and(pred, ~gt.astype(bool)).sum())
    FN = int(np.logical_and(~pred.astype(bool), gt).sum())
    TN = int(np.logical_and(~pred.astype(bool), ~gt.astype(bool)).sum())

    def safe(a, b): return a / b if b else 0.0

    prec = safe(TP, TP + FP)
    rec  = safe(TP, TP + FN)
    f1   = safe(2 * prec * rec, prec + rec)
    iou  = safe(TP, TP + FP + FN)
    dice = safe(2 * TP, 2 * TP + FP + FN)
    return {"TP": TP, "FP": FP, "FN": FN, "TN": TN,
            "precision": prec, "recall": rec,
            "f1": f1, "iou": iou, "dice": dice}


def trapz_auc(values):
    n = len(values)
    if n < 2: return float(values[0]) if n else 0.0
    xs = np.linspace(0, 1, n)
    return float(np.trapz(values, xs))


# ---------------------------------------------------------------------------
# Visualisation
# ---------------------------------------------------------------------------

def _viridis(arr_01: np.ndarray) -> np.ndarray:
    pts = np.array([
        [ 68,  1, 84], [ 72, 35,116], [ 64, 67,135], [ 52, 95,141],
        [ 41,120,142], [ 32,143,140], [ 34,167,132], [ 57,189,117],
        [ 96,208, 97], [143,223, 68], [188,233, 45], [228,239, 35],
        [253,231, 37], [253,231, 37], [253,231, 37], [253,231, 37],
    ], dtype=np.float32)
    t  = np.linspace(0, 1, len(pts))
    t2 = np.linspace(0, 1, 256)
    lut = np.stack([np.interp(t2, t, pts[:, c]) for c in range(3)], axis=-1).astype(np.uint8)
    idx = (np.clip(arr_01, 0, 1) * 255).astype(np.uint8)
    return lut[idx]


def _draw_label_row(width: int, labels: list, height: int = 36) -> np.ndarray:
    _FONT5x7 = {
        ' ': [0x00,0x00,0x00,0x00,0x00],
        'A': [0x7e,0x09,0x09,0x09,0x7e], 'B': [0x7f,0x49,0x49,0x49,0x36],
        'C': [0x3e,0x41,0x41,0x41,0x22], 'D': [0x7f,0x41,0x41,0x41,0x3e],
        'E': [0x7f,0x49,0x49,0x49,0x41], 'F': [0x7f,0x09,0x09,0x09,0x01],
        'G': [0x3e,0x41,0x41,0x49,0x7a], 'H': [0x7f,0x08,0x08,0x08,0x7f],
        'I': [0x00,0x41,0x7f,0x41,0x00], 'J': [0x20,0x40,0x41,0x3f,0x01],
        'K': [0x7f,0x08,0x14,0x22,0x41], 'L': [0x7f,0x40,0x40,0x40,0x40],
        'M': [0x7f,0x02,0x04,0x02,0x7f], 'N': [0x7f,0x04,0x08,0x10,0x7f],
        'O': [0x3e,0x41,0x41,0x41,0x3e], 'P': [0x7f,0x09,0x09,0x09,0x06],
        'Q': [0x3e,0x41,0x51,0x21,0x5e], 'R': [0x7f,0x09,0x19,0x29,0x46],
        'S': [0x46,0x49,0x49,0x49,0x31], 'T': [0x01,0x01,0x7f,0x01,0x01],
        'U': [0x3f,0x40,0x40,0x40,0x3f], 'V': [0x1f,0x20,0x40,0x20,0x1f],
        'W': [0x3f,0x40,0x30,0x40,0x3f], 'X': [0x63,0x14,0x08,0x14,0x63],
        'Y': [0x07,0x08,0x70,0x08,0x07], 'Z': [0x61,0x51,0x49,0x45,0x43],
        'a': [0x20,0x54,0x54,0x54,0x78], 'b': [0x7f,0x48,0x44,0x44,0x38],
        'c': [0x38,0x44,0x44,0x44,0x20], 'd': [0x38,0x44,0x44,0x48,0x7f],
        'e': [0x38,0x54,0x54,0x54,0x18], 'f': [0x08,0x7e,0x09,0x01,0x02],
        'g': [0x0c,0x52,0x52,0x52,0x3e], 'h': [0x7f,0x08,0x04,0x04,0x78],
        'i': [0x00,0x44,0x7d,0x40,0x00], 'j': [0x20,0x40,0x44,0x3d,0x00],
        'k': [0x7f,0x10,0x28,0x44,0x00], 'l': [0x00,0x41,0x7f,0x40,0x00],
        'm': [0x7c,0x04,0x18,0x04,0x78], 'n': [0x7c,0x08,0x04,0x04,0x78],
        'o': [0x38,0x44,0x44,0x44,0x38], 'p': [0x7c,0x14,0x14,0x14,0x08],
        'q': [0x08,0x14,0x14,0x18,0x7c], 'r': [0x7c,0x08,0x04,0x04,0x08],
        's': [0x48,0x54,0x54,0x54,0x20], 't': [0x04,0x3f,0x44,0x40,0x20],
        'u': [0x3c,0x40,0x40,0x20,0x7c], 'v': [0x1c,0x20,0x40,0x20,0x1c],
        'w': [0x3c,0x40,0x30,0x40,0x3c], 'x': [0x44,0x28,0x10,0x28,0x44],
        'y': [0x0c,0x50,0x50,0x50,0x3c], 'z': [0x44,0x64,0x54,0x4c,0x44],
        '+': [0x08,0x08,0x3e,0x08,0x08], '/': [0x20,0x10,0x08,0x04,0x02],
        '(': [0x00,0x1c,0x22,0x41,0x00], ')': [0x00,0x41,0x22,0x1c,0x00],
        '-': [0x08,0x08,0x08,0x08,0x08], '_': [0x40,0x40,0x40,0x40,0x40],
        ':': [0x00,0x36,0x36,0x00,0x00], '.': [0x00,0x60,0x60,0x00,0x00],
        ',': [0x00,0x50,0x30,0x00,0x00],
        '1': [0x00,0x42,0x7f,0x40,0x00], '2': [0x42,0x61,0x51,0x49,0x46],
        '3': [0x21,0x41,0x45,0x4b,0x31], '4': [0x18,0x14,0x12,0x7f,0x10],
        '5': [0x27,0x45,0x45,0x45,0x39], '6': [0x3c,0x4a,0x49,0x49,0x30],
        '7': [0x01,0x71,0x09,0x05,0x03], '8': [0x36,0x49,0x49,0x49,0x36],
        '9': [0x06,0x49,0x49,0x29,0x1e], '0': [0x3e,0x51,0x49,0x45,0x3e],
    }
    CHAR_W, CHAR_H, CHAR_GAP = 5, 7, 1
    SCALE = 2

    def _render_text(text: str, col_w: int) -> np.ndarray:
        strip = np.ones((height, col_w, 3), dtype=np.uint8) * 40
        chars = [_FONT5x7.get(c, _FONT5x7[' ']) for c in text]
        text_w = (len(chars) * (CHAR_W + CHAR_GAP) - CHAR_GAP) * SCALE
        x0 = max(0, (col_w - text_w) // 2)
        y0 = max(0, (height - CHAR_H * SCALE) // 2)
        for ci, cols_bits in enumerate(chars):
            for xi, col_bits in enumerate(cols_bits):
                for yi in range(CHAR_H):
                    if col_bits & (1 << yi):
                        px = x0 + (ci * (CHAR_W + CHAR_GAP) + xi) * SCALE
                        py = y0 + yi * SCALE
                        for dy in range(SCALE):
                            for dx in range(SCALE):
                                if 0 <= py+dy < height and 0 <= px+dx < col_w:
                                    strip[py+dy, px+dx] = [255, 255, 220]
        return strip

    n = len(labels)
    col_w = width // n
    strips = [_render_text(lbl, col_w) for lbl in labels]
    row = np.concatenate(strips, axis=1)
    if row.shape[1] < width:
        pad = np.ones((height, width - row.shape[1], 3), dtype=np.uint8) * 40
        row = np.concatenate([row, pad], axis=1)
    return row[:, :width]


def make_vis(rgb_t, rgb_t1, heatmap, pred, gt, frame_idx: int = None,
             is_event: bool = False, is_pred_event: bool = False):
    H, W = rgb_t.shape[:2]
    c1 = rgb_t.copy()
    c2 = rgb_t1.copy()
    mn, mx = heatmap.min(), heatmap.max()
    c3 = _viridis((heatmap - mn) / (mx - mn + 1e-8))
    c4 = np.stack([pred * 255] * 3, axis=-1).astype(np.uint8)
    c5 = np.stack([gt   * 255] * 3, axis=-1).astype(np.uint8)
    c6 = rgb_t1.copy()
    c6[np.logical_and(pred == 1, gt == 1)] = [  0, 200,   0]
    c6[np.logical_and(pred == 1, gt == 0)] = [220,   0,   0]
    c6[np.logical_and(pred == 0, gt == 1)] = [  0,   0, 220]

    grid = np.concatenate([c1, c2, c3, c4, c5, c6], axis=1)
    total_w = grid.shape[1]

    labels = [
        "Frame t (input)", "Frame t+1 (input)", "Error heatmap",
        "Pred mask (white=embolism)", "GT mask (white=embolism)", "Overlay: G=TP R=FP B=FN",
    ]
    label_row = _draw_label_row(total_w, labels, height=36)
    tags = []
    if is_event:      tags.append("GT EVENT")
    if is_pred_event: tags.append("PRED EVENT")
    event_tag = ("  [" + " + ".join(tags) + "]") if tags else ""
    idx_tag   = f"  Frame pair {frame_idx}" if frame_idx is not None else ""
    info_row  = _draw_label_row(total_w, [f"{idx_tag}{event_tag}"], height=28)

    return np.concatenate([info_row, label_row, grid], axis=0)


# ---------------------------------------------------------------------------
# Video export
# ---------------------------------------------------------------------------

def _make_video(vis_dir: str, out_dir: str, fps: int = 4) -> None:
    import cv2 as _cv2
    frames = sorted(f for f in os.listdir(vis_dir) if f.endswith(".png"))
    if not frames:
        return
    first = _cv2.imread(os.path.join(vis_dir, frames[0]))
    if first is None:
        return
    H, W = first.shape[:2]
    video_path = os.path.join(out_dir, "visualisation.mp4")
    writer = _cv2.VideoWriter(video_path, _cv2.VideoWriter_fourcc(*"mp4v"), fps, (W, H))
    for fname in frames:
        img = _cv2.imread(os.path.join(vis_dir, fname))
        if img is None: continue
        if img.shape[:2] != (H, W):
            img = _cv2.resize(img, (W, H))
        writer.write(img)
    writer.release()
    print(f"  Video ({len(frames)} frames, {fps} fps) → {video_path}")


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def _get_logger(log_path: str) -> logging.Logger:
    logger = logging.getLogger("method.evaluate")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fh = logging.FileHandler(log_path, mode="a")
    fh.setFormatter(logging.Formatter("%(asctime)s  %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(fh)
    logger.addHandler(sh)
    return logger


def _fp50_frame(cum_area: list) -> int:
    total = cum_area[-1] if cum_area else 0
    if total == 0: return -1
    threshold = total * 0.5
    for i, v in enumerate(cum_area):
        if v >= threshold: return i
    return len(cum_area) - 1


def _save_discrete_fp50(pred_cum, gt_cum, seq_name, out_path):
    try:
        import matplotlib.pyplot as plt
        frames     = list(range(len(pred_cum)))
        pred_total = pred_cum[-1] if pred_cum else 0
        gt_total   = gt_cum[-1]   if gt_cum   else 0
        pred_pct   = [v / pred_total * 100 if pred_total > 0 else 0 for v in pred_cum]
        gt_pct     = [v / gt_total   * 100 if gt_total   > 0 else 0 for v in gt_cum]
        pred_fp50  = _fp50_frame(pred_cum)
        gt_fp50    = _fp50_frame(gt_cum)

        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(frames, pred_pct, label="Predicted", lw=1.5)
        ax.plot(frames, gt_pct,   label="GT",        lw=1.5, ls="--")
        if pred_fp50 >= 0: ax.axvline(pred_fp50, color="C0", ls=":", lw=1.5, label=f"fp50 pred (frame {pred_fp50})")
        if gt_fp50   >= 0: ax.axvline(gt_fp50,   color="C1", ls=":", lw=1.5, label=f"fp50 GT (frame {gt_fp50})")
        ex_idx, ex_imgno = _excel_fp50_frame(seq_name)
        if ex_idx is not None:
            ax.axvline(ex_idx, color="C2", ls="-.", lw=1.5, label=f"fp50 GT Excel (frame {ex_imgno})")
        ax.set_xlabel("Frame pair index"); ax.set_ylabel("% of embolism")
        ax.set_title(f"Discrete FP50 — {seq_name}"); ax.set_ylim(0, 105)
        ax.legend(); ax.grid(True, alpha=0.3); fig.tight_layout()
        fig.savefig(out_path, dpi=120); plt.close(fig)
    except ImportError:
        pass


def _find_z_tif(seq_path: str):
    """Return path to the Z*.tif reference image, or None if absent."""
    try:
        for fname in os.listdir(seq_path):
            if fname.lower().startswith("z") and fname.lower().endswith(".tif"):
                return os.path.join(seq_path, fname)
    except OSError:
        pass
    return None


def _save_cumulative_mask(pred_union: np.ndarray, gt_union: np.ndarray,
                          seq_path: str, out_path: str, seq_name: str):
    """Save side-by-side: pred cumulative mask | GT cumulative mask | Z*.tif (if present)."""
    try:
        import matplotlib.pyplot as plt

        z_path = _find_z_tif(seq_path)

        panels, titles = [], []

        panels.append((255 - pred_union * 255).astype(np.uint8))
        titles.append("Pred cumulative mask")

        panels.append((255 - gt_union * 255).astype(np.uint8))
        titles.append("GT cumulative mask")

        if z_path is not None:
            z_arr = np.array(Image.open(z_path).convert("L"))
            if z_arr.shape != pred_union.shape:
                z_arr = np.array(
                    Image.fromarray(z_arr).resize(
                        (pred_union.shape[1], pred_union.shape[0]), Image.NEAREST
                    )
                )
            panels.append(z_arr)
            titles.append(os.path.basename(z_path))

        n = len(panels)
        fig, axes = plt.subplots(1, n, figsize=(6 * n, 5))
        if n == 1:
            axes = [axes]
        for ax, img, title in zip(axes, panels, titles):
            ax.imshow(img, cmap="gray", vmin=0, vmax=255)
            ax.set_title(title, fontsize=9)
            ax.axis("off")
        fig.suptitle(f"Cumulative masks — {seq_name}", fontsize=10)
        fig.tight_layout()
        fig.savefig(out_path, dpi=120)
        plt.close(fig)
    except ImportError:
        pass


# ---------------------------------------------------------------------------
# Per-sequence evaluation
# ---------------------------------------------------------------------------

def evaluate_sequence(seq_name: str, vis_every: int = 10, log=None) -> dict:
    print(f"\n{'='*60}\nEvaluating: {seq_name}\n{'='*60}")

    seq_path = os.path.join(config.DATA_ROOT, seq_name)
    out_dir  = os.path.join(config.OUTPUT_DIR, seq_name)
    eval_dir = os.path.join(config.EVAL_DIR,   seq_name)
    vis_dir  = os.path.join(eval_dir, "visualisations")
    mask_dir = os.path.join(out_dir,  "masks")
    hmap_dir = os.path.join(out_dir,  "heatmaps")

    if not os.path.isdir(mask_dir):
        print(f"  [SKIP] No predicted masks found: {mask_dir}")
        return {}

    frames     = _list_frames(seq_path)
    gt_paths   = _list_masks(seq_path)
    pred_files = sorted(f for f in os.listdir(mask_dir) if f.endswith(".png"))
    n_pairs    = min(len(pred_files), len(gt_paths), len(frames) - 1)
    print(f"  Evaluating {n_pairs} pairs …")

    os.makedirs(vis_dir, exist_ok=True)
    os.makedirs(eval_dir, exist_ok=True)

    all_metrics   = []
    pred_cum_area = []
    gt_cum_area   = []
    pred_union = np.zeros((config.IMG_HEIGHT, config.IMG_WIDTH), dtype=np.uint8)
    gt_union   = np.zeros((config.IMG_HEIGHT, config.IMG_WIDTH), dtype=np.uint8)

    for i in range(n_pairs):
        pred = (np.array(
            Image.open(os.path.join(mask_dir, pred_files[i])).convert("L")
        ) > 128).astype(np.uint8)
        gt = _load_mask(gt_paths[i], config.IMG_HEIGHT, config.IMG_WIDTH)

        m = pixel_metrics(pred, gt)
        m["frame_index"]  = i
        m["frame_name"]   = pred_files[i]
        m["gt_has_event"] = int(gt.any())
        all_metrics.append(m)

        pred_union = np.maximum(pred_union, pred)
        gt_union   = np.maximum(gt_union,   gt)
        prev_p = pred_cum_area[-1] if pred_cum_area else 0
        prev_g = gt_cum_area[-1]   if gt_cum_area   else 0
        pred_cum_area.append(prev_p + int(pred.sum()))
        gt_cum_area.append(  prev_g + int(gt.sum()))

        is_event     = bool(gt.any())
        is_pred_event = bool(pred.any())
        if i % vis_every == 0 or is_event or is_pred_event:
            rgb_t  = _load_rgb(frames[i],   config.IMG_HEIGHT, config.IMG_WIDTH)
            rgb_t1 = _load_rgb(frames[i+1], config.IMG_HEIGHT, config.IMG_WIDTH)
            if rgb_t  is None: rgb_t  = np.zeros((config.IMG_HEIGHT, config.IMG_WIDTH, 3), np.uint8)
            if rgb_t1 is None: rgb_t1 = np.zeros((config.IMG_HEIGHT, config.IMG_WIDTH, 3), np.uint8)
            stem  = os.path.splitext(pred_files[i])[0]
            hpath = os.path.join(hmap_dir, f"{stem}.npy")
            hmap  = np.load(hpath) if os.path.exists(hpath) else np.zeros_like(pred, dtype=np.float32)
            if is_event and is_pred_event:
                tag = "_GTEVENT_PREDEVENT"
            elif is_event:
                tag = "_GTEVENT"
            elif is_pred_event:
                tag = "_PREDEVENT"
            else:
                tag = ""
            vis = make_vis(rgb_t, rgb_t1, hmap, pred, gt,
                           frame_idx=i, is_event=is_event, is_pred_event=is_pred_event)
            Image.fromarray(vis).save(os.path.join(vis_dir, f"frame_{i:04d}{tag}.png"))

    def _mean(key, subset=None):
        rows = [r for r in all_metrics if subset is None or r["gt_has_event"] == subset]
        return float(np.mean([r[key] for r in rows])) if rows else 0.0

    event_rows = [r for r in all_metrics if r["gt_has_event"]]
    white_rows = [r for r in all_metrics if not r["gt_has_event"]]

    pred_fp50_frame = _fp50_frame(pred_cum_area)
    gt_fp50_frame   = _fp50_frame(gt_cum_area)

    summary_lines = [
        f"Evaluation Summary — {seq_name}",
        "=" * 60,
        f"Total pairs: {n_pairs}  |  Event: {len(event_rows)}  |  All-white: {len(white_rows)}",
        "",
        "─── Overall ───────────────────────────────────────────",
        f"  Mean IoU:       {_mean('iou'):.4f}",
        f"  Mean Dice:      {_mean('dice'):.4f}",
        f"  Mean F1:        {_mean('f1'):.4f}",
        f"  Mean Precision: {_mean('precision'):.4f}",
        f"  Mean Recall:    {_mean('recall'):.4f}",
        "",
        "─── Event frames only ─────────────────────────────────",
        f"  Mean IoU:       {_mean('iou', subset=1):.4f}",
        f"  Mean Dice:      {_mean('dice', subset=1):.4f}",
        f"  Mean F1:        {_mean('f1', subset=1):.4f}",
        f"  Mean Precision: {_mean('precision', subset=1):.4f}",
        f"  Mean Recall:    {_mean('recall', subset=1):.4f}",
        "",
        "─── All-white frames (FP rate) ────────────────────────",
        f"  Mean FP pixels: {np.mean([r['FP'] for r in white_rows]):.1f}" if white_rows else "  (none)",
        "",
        "─── FP50 ───────────────────────────────────────────────",
        f"  Pred fp50 frame: {pred_fp50_frame}" if pred_fp50_frame >= 0 else "  Pred fp50 frame: N/A",
        f"  GT   fp50 frame: {gt_fp50_frame}"   if gt_fp50_frame   >= 0 else "  GT   fp50 frame: N/A",
    ]
    for line in summary_lines:
        print(" ", line)

    txt_path = os.path.join(eval_dir, "summary.txt")
    with open(txt_path, "w") as fh:
        fh.write("\n".join(summary_lines) + "\n")

    csv_path = os.path.join(eval_dir, "per_frame_metrics.csv")
    with open(csv_path, "w") as fh:
        fh.write("frame_index,frame_name,gt_has_event,iou,dice,precision,recall,f1,TP,FP,FN,TN\n")
        for r in all_metrics:
            fh.write(
                f"{r['frame_index']},{r['frame_name']},{r['gt_has_event']},"
                f"{r['iou']:.6f},{r['dice']:.6f},{r['precision']:.6f},"
                f"{r['recall']:.6f},{r['f1']:.6f},"
                f"{r['TP']},{r['FP']},{r['FN']},{r['TN']}\n"
            )

    area_csv = os.path.join(eval_dir, "cumulative_area.csv")
    with open(area_csv, "w") as fh:
        fh.write("frame_index,pred_cumulative_area,gt_cumulative_area\n")
        for i, (p, g) in enumerate(zip(pred_cum_area, gt_cum_area)):
            fh.write(f"{i},{p},{g}\n")

    try:
        import matplotlib.pyplot as plt

        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(pred_cum_area, label="Predicted", lw=1.5)
        ax.plot(gt_cum_area,   label="GT",        lw=1.5, ls="--")
        ax.set_xlabel("Frame pair index"); ax.set_ylabel("Cumulative embolised area (px)")
        ax.set_title(f"Cumulative embolism area — {seq_name}")
        ax.legend(); ax.grid(True, alpha=0.3); fig.tight_layout()
        fig.savefig(os.path.join(eval_dir, "cumulative_area.png"), dpi=120); plt.close(fig)

        _save_discrete_fp50(pred_cum_area, gt_cum_area, seq_name,
                            os.path.join(eval_dir, "discrete_fp50.png"))
        _save_fp50_x_wp(pred_cum_area, gt_cum_area, seq_name,
                        os.path.join(eval_dir, "fp50_x_wp.png"))

        fig, axes = plt.subplots(2, 1, figsize=(10, 7), sharex=True)
        axes[0].plot([r["iou"]  for r in all_metrics], lw=0.8)
        axes[0].set_ylabel("IoU"); axes[0].grid(True, alpha=0.3)
        axes[0].set_title(f"Per-frame metrics — {seq_name}")
        axes[1].plot([r["dice"] for r in all_metrics], lw=0.8, color="orange")
        axes[1].set_ylabel("Dice"); axes[1].grid(True, alpha=0.3)
        axes[1].set_xlabel("Frame pair index")
        for r in event_rows:
            axes[0].axvline(r["frame_index"], color="red", alpha=0.2, lw=0.5)
            axes[1].axvline(r["frame_index"], color="red", alpha=0.2, lw=0.5)
        fig.tight_layout()
        fig.savefig(os.path.join(eval_dir, "per_frame_metrics.png"), dpi=120); plt.close(fig)
        print(f"  Plots saved to {eval_dir}/")
    except ImportError:
        pass

    _save_cumulative_mask(pred_union, gt_union, seq_path,
                          os.path.join(eval_dir, "cumulative_mask.png"), seq_name)

    print(f"  Visualisations → {vis_dir}/")
    _make_video(vis_dir, eval_dir, fps=4)

    return {
        "seq_name":        seq_name,
        "n_pairs":         n_pairs,
        "n_event":         len(event_rows),
        "n_white":         len(white_rows),
        "iou":             _mean("iou"),
        "dice":            _mean("dice"),
        "f1":              _mean("f1"),
        "precision":       _mean("precision"),
        "recall":          _mean("recall"),
        "iou_event":       _mean("iou",       subset=1),
        "dice_event":      _mean("dice",      subset=1),
        "f1_event":        _mean("f1",        subset=1),
        "precision_event": _mean("precision", subset=1),
        "recall_event":    _mean("recall",    subset=1),
        "pred_fp50_frame": pred_fp50_frame,
        "gt_fp50_frame":   gt_fp50_frame,
        "pred_cum_area":   pred_cum_area,
        "gt_cum_area":     gt_cum_area,
    }


# ---------------------------------------------------------------------------
# Aggregate helpers
# ---------------------------------------------------------------------------

def _save_avg_cum_plot(all_pred_cum, all_gt_cum, title, out_path):
    try:
        import matplotlib.pyplot as plt
        if not all_pred_cum: return
        max_len = max(max(len(c) for c in all_pred_cum), max(len(c) for c in all_gt_cum))
        xs = np.linspace(0, 1, max_len)

        def _interp(cum_list):
            return [np.interp(xs, np.linspace(0, 1, len(c)), c) for c in cum_list]

        pred_avg = np.mean(_interp(all_pred_cum), axis=0)
        gt_avg   = np.mean(_interp(all_gt_cum),   axis=0)
        frames   = list(range(max_len))

        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(frames, pred_avg, lw=2, label="Predicted (avg)")
        ax.plot(frames, gt_avg,   lw=2, label="GT (avg)", ls="--")
        for c in _interp(all_pred_cum): ax.plot(frames, c, lw=0.5, alpha=0.2, color="C0")
        for c in _interp(all_gt_cum):  ax.plot(frames, c, lw=0.5, alpha=0.2, color="C1")
        ax.set_xlabel("Frame pair index (interpolated)")
        ax.set_ylabel("Cumulative embolised area (px)")
        ax.set_title(title); ax.legend(); ax.grid(True, alpha=0.3); fig.tight_layout()
        fig.savefig(out_path, dpi=120); plt.close(fig)
    except ImportError:
        pass


def _save_avg_fp50_plot(all_pred_cum, all_gt_cum, title, out_path, seq_names=None):
    try:
        import matplotlib.pyplot as plt
        if not all_pred_cum: return
        max_len = max(max(len(c) for c in all_pred_cum), max(len(c) for c in all_gt_cum))
        xs = np.linspace(0, 1, max_len)

        def _interp_pct(cum_list):
            result = []
            for c in cum_list:
                total = c[-1] if c else 0
                pct = [v / total * 100 if total > 0 else 0 for v in c]
                result.append(np.interp(xs, np.linspace(0, 1, len(pct)), pct))
            return result

        pred_interp = _interp_pct(all_pred_cum)
        gt_interp   = _interp_pct(all_gt_cum)
        pred_avg    = np.mean(pred_interp, axis=0)
        gt_avg      = np.mean(gt_interp,   axis=0)
        frames      = list(range(max_len))

        # Aggregate fp50 = mean of per-sequence fp50 (in normalised [0, max_len-1] space).
        # Computed per-seq then averaged so pred/gt/excel are reduced with the same operator,
        # matching the fp50_x_wp_avg logic.
        def _seq_frac(cum_list):
            fracs = []
            for c in cum_list:
                idx = _fp50_frame(c)
                if idx < 0 or len(c) <= 1: continue
                fracs.append(idx / (len(c) - 1))
            return fracs

        pred_fracs = _seq_frac(all_pred_cum)
        gt_fracs   = _seq_frac(all_gt_cum)
        pred_fp50_avg = float(np.mean(pred_fracs)) * (max_len - 1) if pred_fracs else -1
        gt_fp50_avg   = float(np.mean(gt_fracs))   * (max_len - 1) if gt_fracs   else -1

        excel_avg_pos = None
        if seq_names is not None:
            fracs = []
            for name, gt_c in zip(seq_names, all_gt_cum):
                ex_idx, _ = _excel_fp50_frame(name)
                if ex_idx is None or not gt_c or len(gt_c) <= 1: continue
                fracs.append(ex_idx / (len(gt_c) - 1))
            if fracs:
                excel_avg_pos = float(np.mean(fracs)) * (max_len - 1)

        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(frames, pred_avg, lw=2, label="Predicted (avg)")
        ax.plot(frames, gt_avg,   lw=2, label="GT (avg)", ls="--")
        for c in pred_interp: ax.plot(frames, c, lw=0.5, alpha=0.2, color="C0")
        for c in gt_interp:   ax.plot(frames, c, lw=0.5, alpha=0.2, color="C1")
        if pred_fp50_avg >= 0:
            ax.axvline(pred_fp50_avg, color="C0", ls=":", lw=1.5,
                       label=f"fp50 pred avg (frame {pred_fp50_avg:.0f})")
        if gt_fp50_avg >= 0:
            ax.axvline(gt_fp50_avg, color="C1", ls=":", lw=1.5,
                       label=f"fp50 GT avg (frame {gt_fp50_avg:.0f})")
        if excel_avg_pos is not None:
            ax.axvline(excel_avg_pos, color="C2", ls="-.", lw=1.5,
                       label=f"fp50 GT Excel avg (frame {excel_avg_pos:.0f})")
        ax.set_xlabel("Frame pair index (interpolated)")
        ax.set_ylabel("% of embolism")
        ax.set_title(title); ax.set_ylim(0, 105)
        ax.legend(); ax.grid(True, alpha=0.3); fig.tight_layout()
        fig.savefig(out_path, dpi=120); plt.close(fig)
    except ImportError:
        pass


def _save_avg_fp50_x_wp_plot(all_pred_cum, all_gt_cum, seq_names, title, out_path):
    """Average %embolism vs water potential across sequences. Curves of every sequence
    are interpolated onto a common wp grid (intersection of each sequence's wp range)."""
    try:
        import matplotlib.pyplot as plt
    except ImportError:
        return
    if not seq_names: return

    # Collect (wp_array_increasing, pct_pred, pct_gt) per sequence
    series = []  # list of (wp_sorted_asc, p_sorted, g_sorted, fp50_pred_wp, fp50_gt_wp, fp50_ex_wp, fp50_ex_imgno)
    for name, pc, gc in zip(seq_names, all_pred_cum, all_gt_cum):
        imgs, wps = _excel_wp_series(name)
        if imgs is None or not pc or not gc: continue
        img_to_wp = dict(zip(imgs, wps))
        n = min(len(pc), len(gc))
        ptot = pc[-1] if pc else 0; gtot = gc[-1] if gc else 0
        rows = []
        for i in range(n):
            wp = img_to_wp.get(i + 1)
            if wp is None: continue
            rows.append((wp,
                         pc[i] / ptot * 100 if ptot > 0 else 0,
                         gc[i] / gtot * 100 if gtot > 0 else 0))
        if len(rows) < 2: continue
        rows.sort(key=lambda r: r[0])  # ascending wp (most negative first)
        wp_arr = np.array([r[0] for r in rows])
        p_arr  = np.array([r[1] for r in rows])
        g_arr  = np.array([r[2] for r in rows])
        pred_i = _fp50_frame(pc); gt_i = _fp50_frame(gc)
        ex_i, ex_imgno = _excel_fp50_frame(name)
        fp50_pred_wp = img_to_wp.get(pred_i + 1) if pred_i >= 0 else None
        fp50_gt_wp   = img_to_wp.get(gt_i   + 1) if gt_i   >= 0 else None
        fp50_ex_wp   = img_to_wp.get(ex_i   + 1) if ex_i is not None else None
        series.append((wp_arr, p_arr, g_arr, fp50_pred_wp, fp50_gt_wp, fp50_ex_wp, ex_imgno))
    if not series: return

    # Common wp grid = intersection of [min, max] across sequences
    wp_lo = max(s[0].min() for s in series)
    wp_hi = min(s[0].max() for s in series)
    if not (wp_lo < wp_hi):
        return
    grid = np.linspace(wp_lo, wp_hi, 200)

    pred_curves, gt_curves = [], []
    for wp_arr, p_arr, g_arr, *_ in series:
        pred_curves.append(np.interp(grid, wp_arr, p_arr))
        gt_curves.append(  np.interp(grid, wp_arr, g_arr))
    pred_avg = np.mean(pred_curves, axis=0)
    gt_avg   = np.mean(gt_curves,   axis=0)

    def _mean(vals):
        v = [x for x in vals if x is not None]
        return float(np.mean(v)) if v else None
    avg_pred_wp = _mean([s[3] for s in series])
    avg_gt_wp   = _mean([s[4] for s in series])
    avg_ex_wp   = _mean([s[5] for s in series])
    avg_ex_img  = _mean([s[6] for s in series])

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(grid, pred_avg, lw=2, label="Predicted (avg)")
    ax.plot(grid, gt_avg,   lw=2, label="GT mask (avg)", ls="--")
    for c in pred_curves: ax.plot(grid, c, lw=0.5, alpha=0.2, color="C0")
    for c in gt_curves:   ax.plot(grid, c, lw=0.5, alpha=0.2, color="C1")

    def _vline(wp, color, ls, label):
        if wp is None: return
        ax.axvline(wp, color=color, ls=ls, lw=1.5, label=label)
        ax.text(wp, 102, f"{wp:.2f} MPa", color=color, rotation=90,
                ha="right", va="top", fontsize=8)

    if avg_pred_wp is not None:
        _vline(avg_pred_wp, "C0", ":",  f"fp50 pred avg (wp={avg_pred_wp:.2f} MPa)")
    if avg_gt_wp is not None:
        _vline(avg_gt_wp,   "C1", ":",  f"fp50 GT mask avg (wp={avg_gt_wp:.2f} MPa)")
    if avg_ex_wp is not None:
        lbl = f"fp50 GT Excel avg (wp={avg_ex_wp:.2f} MPa"
        if avg_ex_img is not None: lbl += f", image_no {avg_ex_img:.1f}"
        lbl += ")"
        _vline(avg_ex_wp,   "C2", "-.", lbl)

    ax.invert_xaxis()
    ax.set_xlabel("Water potential, wp corrected (MPa)")
    ax.set_ylabel("% of embolism")
    ax.set_title(title); ax.set_ylim(0, 105)
    ax.legend(loc="upper left", fontsize=8); ax.grid(True, alpha=0.3); fig.tight_layout()
    fig.savefig(out_path, dpi=120); plt.close(fig)


def _write_aggregate(results: list, eval_dir: str, log=None):
    import csv

    def _avg(key):
        vals = [r[key] for r in results if isinstance(r.get(key), (int, float))]
        return float(np.mean(vals)) if vals else 0.0

    csv_path = os.path.join(eval_dir, "summary_avg.csv")
    fieldnames = ["seq_name", "n_pairs", "n_event", "n_white",
                  "iou", "dice", "f1", "precision", "recall",
                  "iou_event", "dice_event", "f1_event", "precision_event", "recall_event",
                  "pred_fp50_frame", "gt_fp50_frame"]
    with open(csv_path, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in results:
            w.writerow(r)

    txt_lines = [
        "Aggregate Evaluation Summary",
        "=" * 60,
        f"Sequences evaluated: {len(results)}",
        "",
        "─── Overall (mean across sequences) ───────────────────",
        f"  Mean IoU:           {_avg('iou'):.4f}",
        f"  Mean Dice:          {_avg('dice'):.4f}",
        f"  Mean F1:            {_avg('f1'):.4f}",
        f"  Mean Precision:     {_avg('precision'):.4f}",
        f"  Mean Recall:        {_avg('recall'):.4f}",
        "",
        "─── Event frames only (mean across sequences) ─────────",
        f"  Mean IoU:           {_avg('iou_event'):.4f}",
        f"  Mean Dice:          {_avg('dice_event'):.4f}",
        f"  Mean F1:            {_avg('f1_event'):.4f}",
        f"  Mean Precision:     {_avg('precision_event'):.4f}",
        f"  Mean Recall:        {_avg('recall_event'):.4f}",
        "",
        "─── FP50 (mean frame index, -1 = no embolism) ─────────",
        f"  Mean pred fp50 frame: {_avg('pred_fp50_frame'):.1f}",
        f"  Mean GT   fp50 frame: {_avg('gt_fp50_frame'):.1f}",
    ]
    txt_path = os.path.join(eval_dir, "summary_avg.txt")
    with open(txt_path, "w") as fh:
        fh.write("\n".join(txt_lines) + "\n")
    for line in txt_lines:
        print(" ", line)

    _save_avg_cum_plot(
        [r["pred_cum_area"] for r in results],
        [r["gt_cum_area"]   for r in results],
        title="Avg cumulative embolism area",
        out_path=os.path.join(eval_dir, "cumulative_area_avg.png"),
    )
    _save_avg_fp50_plot(
        [r["pred_cum_area"] for r in results],
        [r["gt_cum_area"]   for r in results],
        title="Avg discrete FP50",
        out_path=os.path.join(eval_dir, "discrete_fp50_avg.png"),
        seq_names=[r["seq_name"] for r in results],
    )
    _save_avg_fp50_x_wp_plot(
        [r["pred_cum_area"] for r in results],
        [r["gt_cum_area"]   for r in results],
        seq_names=[r["seq_name"] for r in results],
        title="Avg FP50 vs water potential",
        out_path=os.path.join(eval_dir, "fp50_x_wp_avg.png"),
    )

    if log:
        log.info(f"\nAggregate summary → {txt_path}")
        log.info(f"Summary CSV       → {csv_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser()
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--sequence", "-s")
    g.add_argument("--all", "-a", action="store_true")
    p.add_argument("--vis-every", type=int, default=10)
    return p.parse_args()


def main():
    args = parse_args()
    os.makedirs(config.EVAL_DIR, exist_ok=True)
    log = _get_logger(os.path.join(config.EVAL_DIR, "evaluate.log"))

    if args.all:
        seqs = find_sequences(config.DATA_ROOT)
        results = []
        for name, _ in seqs:
            r = evaluate_sequence(name, vis_every=args.vis_every, log=log)
            if r:
                results.append(r)
        if results:
            _write_aggregate(results, config.EVAL_DIR, log)
    else:
        evaluate_sequence(args.sequence, vis_every=args.vis_every, log=log)


if __name__ == "__main__":
    main()
